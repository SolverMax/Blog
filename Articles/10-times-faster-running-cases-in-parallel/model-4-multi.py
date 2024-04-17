# Racks and shelves Model 4, running cases in parallel using multiprocessing

# Dependencies
import pyomo.environ as pyo
from datetime import datetime
import os as os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
import multiprocessing as mp

# User selections
dataFile = os.path.join(os.getcwd() + '\data', 'pallets-20000.xlsx')
casesFile = os.path.join(os.getcwd() + '\data', 'ShelfCases-6-0.xlsx')
dataWorksheet = 'Data'
casesWorksheet = 'Data'
weightedObj = True
maxProcesses = 4

# Solver options
solverName = 'appsi_highs'
verbose = False
loadSolution = True
timeLimit = 1*3600   # per case

# Derived globals
processesToUse = min(maxProcesses, mp.cpu_count())

# Fixed
modelName = 'Racks and shelves - Model 4 parallel multiprocessing'

# Generic loader from Excel file, given worksheet and named range
def LoadFromExcel(excelFile, worksheet, rangeName):
    wb = load_workbook(filename=excelFile, read_only=True)
    ws = wb[worksheet]
    dests = wb.defined_names[rangeName].destinations
    for title, coord in dests:
        min_col, min_row, max_col, max_row = range_boundaries(coord)
        data = ws.iter_rows(min_row, max_row, min_col, max_col, values_only = True)
    df = pd.DataFrame(data)
    return df

# Load data from Excel file
def GetData(dataFile, dataWorksheet, casesFile, casesWorksheet):
    Pallets = LoadFromExcel(dataFile, dataWorksheet, 'Heights')   # List of all pallet sizes (decimetres) that we need to store os shelves
    Pallets.columns = ['Sizes']   # Name the pallets column
    MaxShelves = LoadFromExcel(dataFile, dataWorksheet, 'MaxShelves')   # Maximum number of shelves in a rack
    WeightRacks = LoadFromExcel(dataFile, dataWorksheet, 'WeightRacks')   # Objective function weight on the variable for number of racks
    WeightShelves = LoadFromExcel(dataFile, dataWorksheet, 'WeightShelves')   # Objective function weight on the variable for number of shelves
    ShelfHeights = LoadFromExcel(casesFile, casesWorksheet, 'Cases')   # Height of each shelf in each case
    ShelfHeights.columns = [str(n).zfill(1) for n in range(1, ShelfHeights.shape[1] + 1)]   # Label columns from '1' to 'n'
    PalletsPerShelf = LoadFromExcel(dataFile, dataWorksheet, 'PalletsPerShelf')   # Number of pallets on each shelf
    return Pallets, MaxShelves, ShelfHeights, WeightRacks, WeightShelves, PalletsPerShelf

# Define model data, assigning all data to the Model for the given case
def DefineModelData(Model, case, Pallets, MaxShelves, ShelfHeights, PalletsPerShelf, WeightRacks, WeightShelves):
    Model.MaxShelves = MaxShelves[0].item()
    Model.WeightRacks = WeightRacks[0].item()
    Model.WeightShelves = WeightShelves[0].item()
    Model.PalletsPerShelf = PalletsPerShelf[0].item()
    Model.P = pyo.Set(initialize = range(0, len(Pallets)))
    Model.S = pyo.Set(initialize = range(0, Model.MaxShelves))
    Model.Pallets = pyo.Param(Model.P, within = pyo.NonNegativeIntegers, mutable = True)
    Model.ShelfHeights = pyo.Param(Model.S, within = pyo.NonNegativeIntegers, mutable = True)
    Model.NumShelves = 0

    for p in Model.P:
        Model.Pallets[p] = Pallets['Sizes'][p]
    for s in Model.S:
        Model.ShelfHeights[s] = ShelfHeights[str(s+1)][case]
        if ShelfHeights[str(s+1)][case] > 0:
            Model.NumShelves += 1
    return Model
    
# Define optimization model
def defineModel(Model):
    Model.Racks = pyo.Var(domain = pyo.NonNegativeIntegers, initialize = 0)   # Number of racks
    Model.Allocation = pyo.Var(Model.P, Model.S, domain = pyo.Binary, initialize = 0)   # Allocation of pallets to shelves and racks

    def rule_fit(Model, P):   # Each pallet must be allocated to a shelf that is at least the height of the pallet
        return sum(Model.ShelfHeights[s] * Model.Allocation[P, s] for s in Model.S) >= Model.Pallets[P]
    Model.PalletFits = pyo.Constraint(Model.P, rule = rule_fit)

    def rule_use(Model, P):   # Each pallet must be allocated to exactly one shelf
        return sum(Model.Allocation[P, s] for s in Model.S) == 1
    Model.MustUse = pyo.Constraint(Model.P, rule = rule_use)

    def rule_within(Model, S):   # Times each shelf size is allocated to a pallet must be no larger than the number of racks
        return sum(Model.Allocation[p, S] for p in Model.P) <= Model.Racks * Model.PalletsPerShelf   # Some shelves may be empty (Allocation = 0)
    Model.WithinRack = pyo.Constraint(Model.S, rule = rule_within)
    
    def rule_Obj(Model):   # Minimize the number of racks we need to allocate all pallets to a shelf
        if weightedObj:   # Weighted to also minimize number of shelves in a rack, if required
            return Model.WeightRacks * Model.Racks + Model.WeightShelves * Model.NumShelves
        else:
            return Model.Racks
    Model.Obj = pyo.Objective(rule = rule_Obj, sense = pyo.minimize)
    return Model

# Solve model
def solveModel(Model):
    Solver = pyo.SolverFactory(solverName)
    Solver.options['time_limit'] = timeLimit
    Results = Solver.solve(Model, load_solutions = loadSolution, tee = verbose)
    return Model, Results

# One-off setup tasks
def setup(processes):
    print(f'Running {processes} processes on {mp.cpu_count()} cores/threads\n')   # Report number of processes. cpu_count = number of cpu threads
    
# Tasks to be completed
def tasks(case):
    print(f'Start task {case + 1} at {datetime.now().strftime("%H:%M:%S")}')
    Model = pyo.ConcreteModel(name = modelName + ', Case ' + str(case + 1))
    Pallets, MaxShelves, ShelfHeights, WeightRacks, WeightShelves, PalletsPerShelf = GetData(dataFile, dataWorksheet, casesFile, casesWorksheet)
    DefineModelData(Model, case, Pallets, MaxShelves, ShelfHeights, PalletsPerShelf, WeightRacks, WeightShelves)
    defineModel(Model)
    Model, Results = solveModel(Model)
    Obj = round(pyo.value(Model.Obj()), 4)
    Solution = f'Case {case+1:3,.0f}, shelves = '
    for s in Model.S:
        Solution += f'{pyo.value(Model.ShelfHeights[s]):3,.0f}, '
    Solution += f'obj = {Model.Obj():7,.1f}'
    
    print(f'Complete task {case + 1} at {datetime.now().strftime("%H:%M:%S")}')
    return Obj, Solution

# Output summary of case results
def summaryOutput(result):
    BestObj = np.inf
    BestCase = ''
    print()
    for r in range(0, len(result)):
        print(result[r][1])
        currObj = round(result[r][0], 4)
        if currObj < BestObj:
            BestObj = currObj
            BestCase = result[r][1]
    print(f'\nBest case: {BestCase}')



# Main
if __name__ == '__main__':
    dataRows = LoadFromExcel(casesFile, casesWorksheet, 'Cases')   # Get number of rack designs
    numRows = len(dataRows)
    print(f'Number of cases: {numRows}')
    cases = range(0, numRows)   # Define each rack design as a case
    pool = mp.Pool(processes = processesToUse)   # Create a pool with the user-input number of processes
    pool.map(func = setup, iterable = [processesToUse])   # Do one-off setup
    result = pool.map(func = tasks, iterable = cases)   # Run each of the tasks. Returned structure is in case order, irrespective of completion order
    summaryOutput(result)