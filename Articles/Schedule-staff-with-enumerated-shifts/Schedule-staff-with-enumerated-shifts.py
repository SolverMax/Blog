# Schedule staff with enumerated shifts

# Import dependencies
import pyomo.environ as pyo
import pandas as pd
import os.path
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

# Globals
Model = pyo.ConcreteModel(name = 'Schedule with enumerated shifts')
ExcelFile = 'Schedule-staff-with-enumerated-shifts.xlsx'
Worksheet = 'Model 1'
Engine = 'cbc'   # cbc, glpk, or appsi_highs
TimeLimit = 60   # seconds

# Generic loader from Excel file, given worksheet and named range
def LoadFromExcel(ExcelFile, Worksheet, Range):
    Workbook = load_workbook(filename=ExcelFile, read_only = True)
    Sheet = Workbook[Worksheet]
    NamesRanges = Workbook.defined_names[Range].destinations
    for Title, Coord in NamesRanges:
        MinCol, MinRow, MaxCol, MaxRow = range_boundaries(Coord)
        Data = Sheet.iter_rows(MinRow, MaxRow, MinCol, MaxCol, values_only = True)
    ExtractedData = pd.DataFrame(Data)
    return ExtractedData

# Load data from Excel file
def GetData():
    Available = LoadFromExcel(ExcelFile, Worksheet, 'dAvailable')
    People = LoadFromExcel(ExcelFile, Worksheet, 'dPeople')
    PeopleHourly = LoadFromExcel(ExcelFile, Worksheet, 'dPeopleHourly')
    PeopleMin = LoadFromExcel(ExcelFile, Worksheet, 'dPeopleMin')
    SlotRequired = LoadFromExcel(ExcelFile, Worksheet, 'dPeopleRequired')
    Shifts = LoadFromExcel(ExcelFile, Worksheet, 'dShifts')
    ShiftSlots = LoadFromExcel(ExcelFile, Worksheet, 'dShiftSlots')
    Slots = LoadFromExcel(ExcelFile, Worksheet, 'dSlots')
    SurplusMax = LoadFromExcel(ExcelFile, Worksheet, 'dSurplusMax')
    ShiftRequired = LoadFromExcel(ExcelFile, Worksheet, 'dShiftRequired')
    TimePeriod = LoadFromExcel(ExcelFile, Worksheet, 'dTimePeriod')
    return Available, People, PeopleHourly, PeopleMin, SlotRequired, Shifts, ShiftSlots, Slots, SurplusMax, ShiftRequired, TimePeriod

# Declare model components and initialize data structures
def DefineData():
    Available, People, PeopleHourly, PeopleMin, SlotRequired, Shifts, ShiftSlots, Slots, SurplusMax, ShiftRequired, TimePeriod = GetData()
    
    # Single parameter values
    Model.TimePeriod = pyo.Param(within = pyo.NonNegativeReals, initialize = TimePeriod[0][0].item())   # Hours per time slot
    Model.MaxSurplus = pyo.Param(within = pyo.Any, initialize = SurplusMax[0][0].item())   # Maximum surplus people (all slots)
    Model.Engine = pyo.Param(within = pyo.Any, initialize = Engine)   # Name of solver engine (global)
    Model.TimeLimit = pyo.Param(within = pyo.NonNegativeReals, initialize = TimeLimit)   # Solver time limit (global)
    
    # Sets
    Model.People = pyo.Set(initialize = range(1, len(People) + 1))   # Categories of people
    Model.Shifts = pyo.Set(initialize = range(1, len(Shifts) + 1))   # Enumerated shifts
    Model.Slots = pyo.Set(initialize = range(1, len(Slots.columns) + 1))   # Time slots

    # Parameters using sets
    Model.PeopleLabel = pyo.Param(Model.People, within = pyo.Any, mutable = True)   # Label for each person category (unique)
    Model.Available = pyo.Param(Model.People, within = pyo.NonNegativeIntegers, mutable = True)   # Number of each person category available
    Model.Wage = pyo.Param(Model.People, within = pyo.NonNegativeReals, mutable = True)   # Wage rate per person, $/hour
    Model.PeopleMin = pyo.Param(Model.People, within = pyo.NonNegativeIntegers, mutable = True)   # Minimum number of each person category to be used
    for p in Model.People:    
        Model.PeopleLabel[p] = People[0][p - 1]   # Single column
        Model.Available[p] = Available[0][p - 1]
        Model.Wage[p] = PeopleHourly[0][p - 1]
        Model.PeopleMin[p] = PeopleMin[0][p - 1]
        
    Model.ShiftLabel = pyo.Param(Model.Shifts, within = pyo.Any, mutable = True)   # Person category of each shift (non-unique)
    Model.ShiftRequired = pyo.Param(Model.Shifts, within = pyo.NonNegativeIntegers, mutable = True)   # Minimum number of people allocated to each shift
    for s in Model.Shifts:
        Model.ShiftLabel[s] = Shifts[0][s - 1]   # Single column
        Model.ShiftRequired[s] = ShiftRequired[0][s - 1]

    Model.SlotLabel = pyo.Param(Model.Slots, within = pyo.Any, mutable = True)   # Label for each time slot (output only)
    Model.SlotRequired = pyo.Param(Model.Slots, within = pyo.NonNegativeIntegers, mutable = True)   # Number of people required in each time slot
    for t in Model.Slots:
        Model.SlotLabel[t] = Slots[t - 1][0]   # Single row
        Model.SlotRequired[t] = SlotRequired[t - 1][0]

    Model.ShiftSlots = pyo.Param(Model.Shifts, Model.Slots, within = pyo.NonNegativeIntegers, mutable = True)   # Definition of each shift by time slot (0, 1)
    for s in Model.Shifts:
        for t in Model.Slots: 
            Model.ShiftSlots[s, t] = ShiftSlots[t - 1][s - 1] # Translate from 0-base to 1-base
            
# Print the data, if we want to check it
def CheckData(Display):
    if Display:
        print('Data check')
        print('----------\n')
        
        print('Time period: ', pyo.value(Model.TimePeriod))
        print('Max surplus: ', pyo.value(Model.MaxSurplus))
        print('Engine: ', pyo.value(Model.Engine))
        print('TimeLimit: ', pyo.value(Model.TimeLimit))

        print('\nPeople set:', end = ' ')
        for p in Model.People:
            print(p, end = ' ')

        print('\nShifts set:', end = ' ')
        for s in Model.Shifts: 
            print(s, end = ' ')

        print('\nSlots set:', end = ' ')
        for t in Model.Slots: 
            print(t, end = ' ') 

        print('\nPeople labels:', end = ' ')
        for p in Model.People:
            print(pyo.value(Model.PeopleLabel[p]), end = ' ')

        print('\nPeople attributes:')
        for p in Model.People:
            print(f'{pyo.value(Model.PeopleLabel[p])}:  Available {pyo.value(Model.Available[p]):4}  Min {pyo.value(Model.PeopleMin[p]):4}  Wage {pyo.value(Model.Wage[p]):6.2f}')

        print('\nShift labels:', end = ' ')
        for s in Model.Shifts:
            print(pyo.value(Model.ShiftLabel[s]), end = ' ')

        print('\nShift required:', end = ' ')
        for s in Model.Shifts:
            print(pyo.value(Model.ShiftRequired[s]), end = ' ')

        print('\nSlot labels:', end = ' ')
        for t in Model.Slots:
            print(pyo.value(Model.SlotLabel[t]), end = ' ')

        print('\nSlot required:', end = ' ')
        for t in Model.Slots:
            print(pyo.value(Model.SlotRequired[t]), end = ' ')

        print('\nShift slots:')
        for s in Model.Shifts:
            for t in Model.Slots: 
                print(pyo.value(Model.ShiftSlots[s, t]), end = ' ')
            print()
        
        print()
        
# Define model
def DefineModel():
    Model.Allocation = pyo.Var(Model.Shifts, domain = pyo.NonNegativeIntegers)   # Number of each shift to use

    def rule_demand(Model, t):   # Ensure we have the required number of staff, in total, in each time slot
        return sum(Model.ShiftSlots[s, t] * Model.Allocation[s] for s in Model.Shifts) >= Model.SlotRequired[t]
    Model.PeopleRequired = pyo.Constraint(Model.Slots, rule = rule_demand)

    def rule_available(Model, p):   # Number of staff of each category is within their availability
        return sum(Model.Allocation[s] for s in Model.Shifts if pyo.value(Model.ShiftLabel[s]) == pyo.value(Model.PeopleLabel[p])) <= Model.Available[p]
    Model.PeopleAvailable = pyo.Constraint(Model.People, rule = rule_available)

    def rule_peoplemin(Model, p):   # Number of staff of each category is at least the minimum required
        return sum(Model.Allocation[s] for s in Model.Shifts if pyo.value(Model.ShiftLabel[s]) == pyo.value(Model.PeopleLabel[p])) >= Model.PeopleMin[p]
    Model.PeopleLB = pyo.Constraint(Model.People, rule = rule_peoplemin)

    def rule_mustuseshift(Model, s):   # Must use specific shifts, if required
        return Model.Allocation[s] >= Model.ShiftRequired[s]
    Model.ShiftMustUse = pyo.Constraint(Model.Shifts, rule = rule_mustuseshift)

    def rule_surplus(Model, t):   # The number of surplus staff in each time slot is no more than the maximum allowed
        return sum(Model.ShiftSlots[s, t] * Model.Allocation[s] for s in Model.Shifts) - Model.SlotRequired[t] <= Model.MaxSurplus
    Model.SlotSurplus = pyo.Constraint(Model.Slots, rule = rule_surplus)

    # def rule_ObjAlt(Model):   # Alternative, equivalent way of writing the objective function
    #     Cost = 0
    #     for p in Model.People:
    #         Cost += sum(sum(Model.ShiftSlots[s, t] * Model.Allocation[s] for s in Model.Shifts if pyo.value(Model.ShiftLabel[s]) == pyo.value(Model.PeopleLabel[p])) for t in Model.Slots) * pyo.value(Model.Wage[p]) * Model.TimePeriod
    #     return Cost
    # Model.TotalCostAlt = pyo.Objective(rule = rule_ObjAlt, sense = pyo.minimize)

    def rule_Obj(Model):   # Minimize total cost of staff
        Cost = 0
        for p in Model.People:
            for s in Model.Shifts:
                if pyo.value(Model.ShiftLabel[s]) == pyo.value(Model.PeopleLabel[p]):
                    Cost += sum(Model.ShiftSlots[s, t] * Model.Allocation[s] for t in Model.Slots) * pyo.value(Model.Wage[p])
        Cost = Cost * Model.TimePeriod
        return Cost
    Model.TotalCost = pyo.Objective(rule = rule_Obj, sense = pyo.minimize)
    
# Solve model
def SolveModel(Verbose):
    Solver = pyo.SolverFactory(pyo.value(Model.Engine))

    if pyo.value(Model.Engine) == 'cbc':
        Solver.options['seconds'] = pyo.value(Model.TimeLimit)
    elif pyo.value(Model.Engine) == 'glpk':
        Solver.options['tmlim'] = pyo.value(Model.TimeLimit)
    elif pyo.value(Model.Engine) == 'appsi_highs':
        Solver.options['time_limit'] = pyo.value(Model.TimeLimit)

    Results = Solver.solve(Model, load_solutions = False, tee = Verbose)
    return Results

# Process results
def ProcessResults():
    WriteSolution = False
    Optimal = False
    LimitStop = False
    Condition = Results.solver.termination_condition

    if Condition == pyo.TerminationCondition.optimal:
        Optimal = True
    if Condition == pyo.TerminationCondition.maxTimeLimit or Condition == pyo.TerminationCondition.maxIterations:
        LimitStop = True
    if Optimal or LimitStop:
        try:
            WriteSolution = True
            Model.solutions.load_from(Results)   # Defer loading results until now, in case there is no solution to load
            SolverData = Results.Problem._list
            SolutionLB = SolverData[0].lower_bound
            SolutionUB = SolverData[0].upper_bound
        except:
            WriteSolution = False
            LimitStop = False
    return WriteSolution, LimitStop, SolutionLB, SolutionUB

# Write solution
def Output(WriteSolution, LimitStop, SolutionLB, SolutionUB):
    print('Model: ', Model.name)
    print('Status:', Results.solver.termination_condition)
    print('Solver:', pyo.value(Model.Engine), '\n')
    if LimitStop:
        print('Objective bounds')
        print('----------------')
        if SolutionLB is None:
            print('Lower:      None')
        else:
            print(f'Lower: {SolutionLB:9,.2f}')
        if SolutionUB is None:
            print('Upper:      None\n')
        else:
            print(f'Upper: {SolutionUB:9,.2f}\n')
    if WriteSolution:
        print(f'Total cost = ${Model.TotalCost():11,.2f}\n')
        pd.set_option('display.max_rows', None)
        pd.set_option('display.max_columns', None)
        pd.options.display.float_format = '{:.0f}'.format

        Summary = pd.DataFrame()
        for p in Model.People:
            PeopleUsed = sum(pyo.value(Model.Allocation[s]) for s in Model.Shifts if pyo.value(Model.ShiftLabel[s]) == pyo.value(Model.PeopleLabel[p]))
            Summary.loc[pyo.value(Model.PeopleLabel[p]), 'Used'] = PeopleUsed
            Summary.loc[pyo.value(Model.PeopleLabel[p]), 'Available'] = pyo.value(Model.Available[p])
            Summary.loc[pyo.value(Model.PeopleLabel[p]), 'At least'] = pyo.value(Model.PeopleMin[p])
        display(Summary)

        Detail = pd.DataFrame()
        for s in Model.Shifts:
            if pyo.value(Model.Allocation[s]) >= 0.5:   # Use 0.5 to allow for precision issues
                for t in Model.Slots: 
                    Detail.loc[str(s) + '-' + pyo.value(Model.ShiftLabel[s]), str(pyo.value(Model.SlotLabel[t]))[0:5]] = pyo.value(Model.ShiftSlots[s, t]) * pyo.value(Model.Allocation[s])
        for s in Model.Shifts:
            for t in Model.Slots: 
                Detail.loc['Total', str(pyo.value(Model.SlotLabel[t]))[0:5]] = sum(pyo.value(Model.ShiftSlots[s, t] * Model.Allocation[s]) for s in Model.Shifts)
                Detail.loc['Required', str(pyo.value(Model.SlotLabel[t]))[0:5]] = pyo.value(Model.SlotRequired[t])
                Detail.loc['Surplus', str(pyo.value(Model.SlotLabel[t]))[0:5]] = sum(pyo.value(Model.ShiftSlots[s, t] * Model.Allocation[s]) for s in Model.Shifts) - pyo.value(Model.SlotRequired[t])
        display(Detail)
    else:
        print('No solution loaded')
        print('Model:')
        Model.pprint()
        
# Main
DefineData()
CheckData(Display=True)
DefineModel()
Results = SolveModel(Verbose=False)
WriteSolution, LimitStop, SolutionLB, SolutionUB = ProcessResults()
Output(WriteSolution, LimitStop, SolutionLB, SolutionUB)